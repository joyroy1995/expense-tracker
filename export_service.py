import csv
import io
from datetime import datetime

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def _month_name(m):
    return MONTH_NAMES[m - 1] if 1 <= m <= 12 else ""


def generate_csv(expenses, year, month):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Description", "Category", "Amount"])
    for e in expenses:
        writer.writerow([e["date"], e["description"], e["category"], f'{e["amount"]:.2f}'])
    total = sum(e["amount"] for e in expenses)
    writer.writerow([])
    writer.writerow(["Total", "", "", f"{total:.2f}"])
    return output.getvalue()


def generate_xlsx(expenses, year, month):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"{_month_name(month)} {year}"

    header_font = Font(bold=True, color="ffffff", size=11)
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
        top=Side(style="thin", color="e2e8f0"),
        bottom=Side(style="thin", color="e2e8f0"),
    )

    headers = ["Date", "Description", "Category", "Amount (BDT)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for i, e in enumerate(expenses, 2):
        ws.cell(row=i, column=1, value=e["date"]).border = thin_border
        ws.cell(row=i, column=2, value=e["description"]).border = thin_border
        ws.cell(row=i, column=3, value=e["category"]).border = thin_border
        c = ws.cell(row=i, column=4, value=e["amount"])
        c.number_format = '#,##0.00'
        c.border = thin_border

    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin_border
    total_cell = ws.cell(row=total_row, column=4, value=sum(e["amount"] for e in expenses))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'
    total_cell.border = thin_border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf(expenses, year, month):
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(99, 102, 241)
    pdf.cell(0, 12, f"Expense Report - {_month_name(month)} {year}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(6)

    col_widths = [32, 80, 36, 32]
    headers = ["Date", "Description", "Category", "Amount"]
    header_color = (99, 102, 241)
    alt_row_color = (248, 250, 252)

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(*header_color)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    total = 0
    for idx, e in enumerate(expenses):
        total += e["amount"]
        if idx % 2 == 0:
            pdf.set_fill_color(*alt_row_color)
            fill = True
        else:
            fill = False
        pdf.set_text_color(30, 41, 59)
        pdf.cell(col_widths[0], 7, e["date"], border=1, fill=fill)
        pdf.cell(col_widths[1], 7, e["description"][:48], border=1, fill=fill)
        pdf.cell(col_widths[2], 7, e["category"], border=1, align="C", fill=fill)
        pdf.cell(col_widths[3], 7, f'{e["amount"]:.2f}', border=1, align="R", fill=fill)
        pdf.ln()

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(sum(col_widths) - col_widths[3], 7, "Total", border="T", align="R")
    pdf.cell(col_widths[3], 7, f'{total:.2f}', border="T", align="R")

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return output
